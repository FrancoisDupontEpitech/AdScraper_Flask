import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
import json

def generate_excel_from_json(company_data, mode, json_file_path):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws["A1"] = "Company Name"
    ws["B1"] = "Advertisement URLs"
    ws["D1"] = "Date"

    # Set text wrapping for column B
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Count total iterations for progress tracking
    total_iterations = sum(len(urls) for urls in company_data.values())

    # Iterate over data and write to Excel
    row = 2
    current_iteration = 0
    for company, ads in company_data.items():
        for ad in ads:
            current_iteration += 1
            print(f"[{current_iteration}/{total_iterations}] Generating data for {company}...")
            ws[f"A{row}"] = ad["company"]
            ws[f"B{row}"] = ad["link"]
            ws[f"D{row}"] = ad.get("date", "N/A")  # Using .get() to handle missing dates gracefully
            row += 1

    # Determine the Excel file path based on the JSON file path
    excel_file_path = json_file_path.rsplit('.', 1)[0] + '.xlsx'

    print(f"Saving Excel file as: {excel_file_path}")
    wb.save(excel_file_path)
    print(f"Excel file saved as: {excel_file_path}")
